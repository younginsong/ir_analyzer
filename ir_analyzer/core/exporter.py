"""
exporter.py - 피팅 결과 Excel 내보내기
단일 파일 / 전체 스펙트럼 / 배치 결과 모두 지원
"""

from __future__ import annotations

import re
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from core.fitter import FitResult


# 컬러 팔레트
HEADER_COLOR = "2F5496"
SUBHEADER_COLOR = "9DC3E6"
ALT_ROW_COLOR = "EBF3FB"


def _style_header(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal='center')


def _autosize_columns(ws, sample_row_limit: int | None = None, min_width: int = 12):
    """워크시트 컬럼 너비를 샘플 행 기준으로 자동 조정."""
    for col in ws.columns:
        sample = []
        for cell in col:
            if cell.value is None:
                continue
            if sample_row_limit is not None and cell.row > sample_row_limit:
                continue
            sample.append(cell)
        max_len = max((len(str(cell.value)) for cell in sample), default=8)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, min_width)


def _shares_wavenumber_grid(entries, atol: float = 1e-6) -> bool:
    """모든 스펙트럼이 동일한 파수축을 공유하는지 확인."""
    if not entries:
        return False

    base_wn = np.asarray(entries[0].wavenumber)
    for entry in entries[1:]:
        wn = np.asarray(entry.wavenumber)
        if len(wn) != len(base_wn):
            return False
        if not np.allclose(wn, base_wn, atol=atol, rtol=0):
            return False
    return True


def _append_oh_processed_sheet(wb, entries, potentials: dict, spectrum_states: dict) -> int:
    """OH baseline/corrected 스펙트럼을 long-format 으로 저장."""
    if not spectrum_states:
        return 0

    ws = wb.create_sheet("OH Processed")
    headers = [
        "Spectrum",
        "Potential (V)",
        "Wavenumber (cm⁻¹)",
        "Raw",
        "Baseline",
        "Corrected",
        "Source File",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    rows_written = 0
    for entry in entries:
        state = spectrum_states.get(entry.filepath)
        if not state:
            continue

        wn = np.asarray(state.get('wn_crop', []))
        ab_raw = np.asarray(state.get('ab_crop', []))
        baseline = np.asarray(state.get('baseline', []))
        corrected = np.asarray(state.get('ab_corrected', []))
        if len(wn) == 0:
            continue

        potential = potentials.get(entry.name)
        for idx in range(len(wn)):
            ws.append([
                entry.name,
                round(potential, 4) if potential is not None else "",
                round(float(wn[idx]), 4),
                round(float(ab_raw[idx]), 6),
                round(float(baseline[idx]), 6) if idx < len(baseline) else "",
                round(float(corrected[idx]), 6) if idx < len(corrected) else "",
                entry.source_spectrum_path or entry.filepath,
            ])
            rows_written += 1

    _autosize_columns(ws, sample_row_limit=20, min_width=14)
    return rows_written


def _append_co_processed_sheet(wb, entries, potentials: dict,
                               spectrum_states: dict) -> int:
    """CO fixed regions에서 Total-corrected 또는 raw 분석 입력을 저장."""
    if not entries:
        return 0

    ws = wb.create_sheet("CO Processed")
    headers = [
        "Spectrum",
        "Potential (V)",
        "Region",
        "Wavenumber (cm⁻¹)",
        "Raw",
        "Baseline",
        "Analysis Input",
        "Region Min (cm⁻¹)",
        "Region Max (cm⁻¹)",
        "Input Source",
        "Source File",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    rows_written = 0
    for entry in entries:
        potential = potentials.get(entry.name)
        wn_full = np.asarray(entry.wavenumber)
        ab_full = np.asarray(entry.absorbance)
        state = spectrum_states.get(entry.filepath, {})
        state_wn = np.asarray(state.get('wn_crop', []))
        state_raw = np.asarray(state.get('ab_crop', []))
        state_baseline = np.asarray(state.get('baseline', []))
        state_corrected = np.asarray(state.get('ab_corrected', []))

        for region, (lo, hi) in {
            'CO_L': (2000.0, 2100.0),
            'CO_B': (1650.0, 1900.0),
        }.items():
            state_mask = (state_wn >= lo) & (state_wn <= hi)
            state_ready = (
                len(state_wn) > 0
                and len(state_raw) == len(state_wn)
                and len(state_baseline) == len(state_wn)
                and len(state_corrected) == len(state_wn)
                and np.any(state_mask)
            )
            if state_ready:
                wn = state_wn[state_mask]
                ab = state_raw[state_mask]
                baseline = state_baseline[state_mask]
                corrected = state_corrected[state_mask]
                input_source = "Total corrected"
            else:
                mask = (wn_full >= lo) & (wn_full <= hi)
                if not np.any(mask):
                    continue
                wn = wn_full[mask]
                ab = ab_full[mask]
                baseline = np.zeros_like(ab)
                corrected = ab.copy()
                input_source = "Raw"

            for idx in range(len(wn)):
                ws.append([
                    entry.name,
                    round(potential, 4) if potential is not None else "",
                    region,
                    round(float(wn[idx]), 4),
                    round(float(ab[idx]), 6),
                    round(float(baseline[idx]), 6),
                    round(float(corrected[idx]), 6),
                    round(lo, 4),
                    round(hi, 4),
                    input_source,
                    entry.source_spectrum_path or entry.filepath,
                ])
                rows_written += 1

    _autosize_columns(ws, sample_row_limit=20, min_width=14)
    return rows_written


def export_spectra_excel(entries, potentials: dict, filepath: str,
                         spectrum_states: dict | None = None,
                         co_states: dict | None = None) -> dict:
    """
    로드된 스펙트럼 전체를 별도 Excel 파일로 저장.

    - Index 시트: 스펙트럼 메타데이터
    - Raw Matrix / Raw Spectra: 원본 스펙트럼
    - OH Processed: OH baseline / corrected
    - CO Processed: CO 분석에 사용되는 Total corrected 또는 raw 입력
    """
    if not entries:
        raise ValueError("내보낼 스펙트럼이 없습니다.")

    wb = Workbook()
    ws_idx = wb.active
    ws_idx.title = "Index"

    idx_headers = [
        "Spectrum",
        "Potential (V)",
        "Source File",
        "Original Name",
        "Session",
        "Points",
        "Wavenumber Min (cm⁻¹)",
        "Wavenumber Max (cm⁻¹)",
    ]
    ws_idx.append(idx_headers)
    _style_header(ws_idx, 1, len(idx_headers))
    ws_idx.freeze_panes = "A2"

    total_points = 0
    for i, entry in enumerate(entries):
        potential = potentials.get(entry.name)
        wn = np.asarray(entry.wavenumber)
        total_points += len(wn)
        ws_idx.append([
            entry.name,
            round(potential, 4) if potential is not None else "",
            entry.source_spectrum_path or entry.filepath,
            entry.original_name or entry.name,
            entry.source_session_label or "",
            len(wn),
            round(float(wn[0]), 4) if len(wn) else "",
            round(float(wn[-1]), 4) if len(wn) else "",
        ])
        if i % 2 == 1:
            for col in range(1, len(idx_headers) + 1):
                ws_idx.cell(row=ws_idx.max_row, column=col).fill = \
                    PatternFill("solid", fgColor=ALT_ROW_COLOR)

    layout = "matrix" if _shares_wavenumber_grid(entries) else "long"

    if layout == "matrix":
        ws_raw = wb.create_sheet("Raw Matrix")
        headers = ["Wavenumber (cm⁻¹)"] + [entry.name for entry in entries]
        ws_raw.append(headers)
        _style_header(ws_raw, 1, len(headers))
        ws_raw.freeze_panes = "A2"

        base_wn = np.asarray(entries[0].wavenumber)
        absorbance_arrays = [np.asarray(entry.absorbance) for entry in entries]
        for idx in range(len(base_wn)):
            row = [round(float(base_wn[idx]), 4)]
            row.extend(round(float(ab[idx]), 6) for ab in absorbance_arrays)
            ws_raw.append(row)

        _autosize_columns(ws_raw, sample_row_limit=12, min_width=14)
    else:
        ws_raw = wb.create_sheet("Raw Spectra")
        headers = [
            "Spectrum",
            "Potential (V)",
            "Wavenumber (cm⁻¹)",
            "Absorbance",
            "Source File",
        ]
        ws_raw.append(headers)
        _style_header(ws_raw, 1, len(headers))
        ws_raw.freeze_panes = "A2"

        for entry in entries:
            potential = potentials.get(entry.name)
            for wn, ab in zip(entry.wavenumber, entry.absorbance):
                ws_raw.append([
                    entry.name,
                    round(potential, 4) if potential is not None else "",
                    round(float(wn), 4),
                    round(float(ab), 6),
                    entry.source_spectrum_path or entry.filepath,
                ])

        _autosize_columns(ws_raw, sample_row_limit=20, min_width=14)

    oh_points = _append_oh_processed_sheet(wb, entries, potentials, spectrum_states or {})
    co_points = _append_co_processed_sheet(
        wb, entries, potentials, spectrum_states or {})

    _autosize_columns(ws_idx, sample_row_limit=20, min_width=14)
    wb.save(filepath)
    return {
        'filepath': filepath,
        'layout': layout,
        'n_spectra': len(entries),
        'n_points': total_points,
        'oh_points': oh_points,
        'co_points': co_points,
    }


def export_single(wavenumber: np.ndarray,
                  absorbance: np.ndarray,
                  baseline: np.ndarray,
                  fit_result: FitResult,
                  filepath: str,
                  filename: str = ""):
    """단일 파일 피팅 결과를 Excel로 저장"""
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = "IR Spectrum Fitting Results"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A2'] = f"File: {filename}"
    ws_summary['A3'] = f"R² = {fit_result.r_squared:.6f}"
    ws_summary['A4'] = f"χ² = {fit_result.chi_squared:.6e}"

    # 피크 결과 테이블
    headers = ["Peak", "Shape", "Center (cm⁻¹)", "Center Err",
               "FWHM (cm⁻¹)", "Amplitude", "Area", "Area Fraction (%)"]
    ws_summary.append([])
    ws_summary.append(headers)
    _style_header(ws_summary, ws_summary.max_row, len(headers))

    for i, p in enumerate(fit_result.peaks):
        row = [
            f"Peak {i+1}",
            p.shape.capitalize(),
            round(p.center, 2),
            round(p.center_err, 4),
            round(p.fwhm, 2),
            round(p.amplitude, 6),
            round(p.area, 6),
            round(p.area_fraction, 2)
        ]
        ws_summary.append(row)
        if i % 2 == 1:
            for col in range(1, len(headers) + 1):
                ws_summary.cell(row=ws_summary.max_row, column=col).fill = \
                    PatternFill("solid", fgColor=ALT_ROW_COLOR)

    # 컬럼 너비
    col_widths = [8, 12, 16, 12, 14, 14, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws_summary.column_dimensions[ws_summary.cell(1, i).column_letter].width = w

    # ── Sheet 2: Spectral Data ────────────────────────────────
    ws_data = wb.create_sheet("Spectral Data")

    data_headers = ["Wavenumber (cm⁻¹)", "Raw Absorbance", "Baseline",
                    "Corrected", "Fitted"]
    for i, p in enumerate(fit_result.peaks):
        data_headers.append(f"Peak {i+1} ({p.center:.0f} cm⁻¹)")

    ws_data.append(data_headers)
    _style_header(ws_data, 1, len(data_headers))

    corrected = absorbance - baseline
    for j in range(len(wavenumber)):
        row = [
            round(wavenumber[j], 2),
            round(absorbance[j], 6),
            round(baseline[j], 6),
            round(corrected[j], 6),
            round(fit_result.fitted_curve[j], 6) if fit_result.success else "",
        ]
        for curve in fit_result.individual_curves:
            row.append(round(curve[j], 6))
        ws_data.append(row)

    # ── Sheet 3: Residuals ────────────────────────────────────
    ws_res = wb.create_sheet("Residuals")
    ws_res.append(["Wavenumber (cm⁻¹)", "Residual"])
    _style_header(ws_res, 1, 2)
    for j in range(len(wavenumber)):
        ws_res.append([round(wavenumber[j], 2),
                        round(fit_result.residual[j], 8)])

    wb.save(filepath)
    return filepath


def export_all_spectra(entries, spectrum_states: dict, potentials: dict,
                       stark_results: list, filepath: str,
                       sio_ref_area=None):
    """
    모든 스펙트럼 피팅 결과를 퍼텐셜별 시트로 분리하여 저장.
    마지막 시트(Summary)에는 Area Fraction 요약 + Stark Tuning Slopes 수록.

    entries        : [SpectrumEntry]  — 스펙트럼 목록 (순서 유지)
    spectrum_states: {filepath: state_dict}
    potentials     : {filename: V}
    stark_results  : [StarkResult]  (없으면 빈 리스트)
    sio_ref_area   : float 또는 {spectrum_name: area}
    """
    # 피팅된 항목만 추출 (목록 순서 유지)
    fitted = []
    for entry in entries:
        state = spectrum_states.get(entry.filepath)
        if state and state.get('fit_result') and state['fit_result'].success:
            fitted.append((entry, state))

    if not fitted:
        raise ValueError("피팅된 스펙트럼이 없습니다.")

    wb = Workbook()
    wb.remove(wb.active)   # 기본 Sheet 제거

    # ── 퍼텐셜별 시트 ─────────────────────────────────────────
    for entry, state in fitted:
        potential = potentials.get(entry.name)
        if potential is not None:
            sheet_name = f"{potential:.3f}V"
        else:
            sheet_name = re.sub(r'[\\/*?:\[\]]', '_', Path(entry.name).stem)[:28]

        # 중복 시트명 방지
        existing = {ws.title for ws in wb.worksheets}
        base, n = sheet_name, 1
        while sheet_name in existing:
            sheet_name = f"{base[:25]}_{n}"; n += 1

        ws = wb.create_sheet(sheet_name)
        fr  = state['fit_result']
        wn  = state['wn_crop']
        ab_raw = state['ab_crop']
        bl  = state['baseline']
        ab_cor = state['ab_corrected']

        # 파일명 / 전위 / R²
        ws['A1'] = entry.name
        ws['A1'].font = Font(bold=True)
        if potential is not None:
            ws['B1'] = f"Potential: {potential:.3f} V"
        ws['A2'] = f"R² = {fr.r_squared:.6f}"

        # 피크 요약 테이블
        ws.append([])
        pk_headers = ["Peak", "Shape", "Center (cm⁻¹)", "FWHM (cm⁻¹)",
                      "Area", "Area Fraction (%)"]
        ws.append(pk_headers)
        _style_header(ws, ws.max_row, len(pk_headers))
        for i, p in enumerate(fr.peaks):
            ws.append([f"Peak {i+1}", p.shape.capitalize(),
                       round(p.center, 2), round(p.fwhm, 2),
                       round(p.area, 6), round(p.area_fraction, 2)])

        # 스펙트럼 데이터 테이블
        ws.append([])
        data_hdrs = ["Wavenumber (cm⁻¹)", "Raw Absorbance", "Baseline",
                     "Corrected", "Fitted"]
        for i, p in enumerate(fr.peaks):
            data_hdrs.append(f"Peak {i+1} ({p.center:.0f} cm⁻¹)")
        ws.append(data_hdrs)
        _style_header(ws, ws.max_row, len(data_hdrs))

        for j in range(len(wn)):
            row = [round(wn[j], 4), round(ab_raw[j], 6),
                   round(bl[j], 6),  round(ab_cor[j], 6),
                   round(fr.fitted_curve[j], 6)]
            for curve in fr.individual_curves:
                row.append(round(curve[j], 6))
            ws.append(row)

        # 컬럼 너비 (처음 몇 행 기준)
        for col in ws.columns:
            sample = [c for c in col if c.row <= 10 and c.value is not None]
            max_len = max((len(str(c.value)) for c in sample), default=8)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    # ── Summary 시트 (마지막) ─────────────────────────────────
    ws_s = wb.create_sheet("Summary")

    # ① Area Fraction 요약
    ws_s['A1'] = "Area Fraction Summary"
    ws_s['A1'].font = Font(bold=True, size=12)

    max_peaks = max(len(e[1]['fit_result'].peaks) for e in fitted)

    af_headers = ["Spectrum", "Potential (V)", "R²"]
    for i in range(max_peaks):
        af_headers += [f"P{i+1} Center (cm⁻¹)", f"P{i+1} FWHM (cm⁻¹)",
                       f"P{i+1} Area", f"P{i+1} Area (%)"]
    af_headers += ["OH Total Area", "OH Area / Si-O Area"]
    ws_s.append([])
    ws_s.append(af_headers)
    _style_header(ws_s, ws_s.max_row, len(af_headers))

    for k, (entry, state) in enumerate(fitted):
        potential = potentials.get(entry.name)
        fr = state['fit_result']
        row = [entry.name,
               round(potential, 4) if potential is not None else "—",
               round(fr.r_squared, 6)]
        for p in fr.peaks:
            row += [round(p.center, 2), round(p.fwhm, 2),
                    round(p.area, 6), round(p.area_fraction, 2)]
        total_oh = sum(p.area for p in fr.peaks)
        row.append(round(total_oh, 6))
        sio_area = (
            sio_ref_area.get(entry.name)
            if isinstance(sio_ref_area, dict)
            else sio_ref_area
        )
        if sio_area:
            row.append(round(total_oh / sio_area, 6))
        else:
            row.append("—")
        ws_s.append(row)
        if k % 2 == 1:
            for col in range(1, len(af_headers) + 1):
                ws_s.cell(row=ws_s.max_row, column=col).fill = \
                    PatternFill("solid", fgColor=ALT_ROW_COLOR)

    # ② Stark Tuning Slopes
    if stark_results:
        ws_s.append([])
        ws_s.append([])
        r_stark = ws_s.max_row
        ws_s.cell(r_stark, 1).value = "Stark Tuning Slopes"
        ws_s.cell(r_stark, 1).font = Font(bold=True, size=12)

        ws_s.append([])
        stark_hdrs = ["Peak", "Slope (cm⁻¹/V)", "Intercept (cm⁻¹)", "R²", "N Points"]
        ws_s.append(stark_hdrs)
        _style_header(ws_s, ws_s.max_row, len(stark_hdrs))

        for sr in stark_results:
            ws_s.append([f"P{sr.peak_index + 1}",
                         round(sr.slope, 4), round(sr.intercept, 4),
                         round(sr.r_squared, 6), sr.n_points])

        # 피크별 potential vs center 원데이터
        ws_s.append([])
        ws_s.append(["Peak-wise data (Potential vs Center):"])
        ws_s.cell(ws_s.max_row, 1).font = Font(bold=True)

        for sr in stark_results:
            ws_s.append([])
            ws_s.append([f"Peak {sr.peak_index + 1}"])
            ws_s.cell(ws_s.max_row, 1).font = Font(bold=True)
            ws_s.append(["Potential (V)", "Center (cm⁻¹)"])
            _style_header(ws_s, ws_s.max_row, 2)
            for pot, ctr in zip(sr.potentials, sr.centers):
                ws_s.append([round(pot, 4), round(ctr, 4)])

    # 컬럼 너비
    _autosize_columns(ws_s, min_width=14)

    wb.save(filepath)
    return filepath


def export_co_results(co_fit_records: list, co_states: dict,
                      entries, potentials: dict, filepath: str,
                      co_stark_results: list = None):
    """
    CO 분석 결과 Excel 저장.
    - CO Summary 시트: 전위별 CO_L/CO_B center, area, ratio
    - CO deconvolution 시트 (manual fit 사용 시): wn, corrected, fitted, 개별 피크 곡선
    """
    entry_map = {e.name: e for e in (entries or [])}

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "CO Summary"

    headers = ["Spectrum", "Potential (V)",
               "CO_L Center (cm⁻¹)", "CO_L Area",
               "CO_B Center (cm⁻¹)", "CO_B Area",
               "CO_L / CO_B Ratio"]
    ws_sum.append(headers)
    _style_header(ws_sum, 1, len(headers))

    for k, record in enumerate(co_fit_records):
        fname = record['filename']
        pot   = potentials.get(fname)
        co_l  = record.get('CO_L')
        co_b  = record.get('CO_B')

        l_center = round(co_l.peaks[0].center, 2) if (co_l and co_l.success and co_l.peaks) else ""
        l_area   = round(co_l.peaks[0].area, 6)   if (co_l and co_l.success and co_l.peaks) else ""
        b_center = round(co_b.peaks[0].center, 2) if (co_b and co_b.success and co_b.peaks) else ""
        b_area   = round(co_b.peaks[0].area, 6)   if (co_b and co_b.success and co_b.peaks) else ""
        ratio    = round(l_area / b_area, 4) if (isinstance(l_area, float) and
                                                  isinstance(b_area, float) and b_area > 0) else ""

        row = [fname,
               round(pot, 4) if pot is not None else "",
               l_center, l_area, b_center, b_area, ratio]
        ws_sum.append(row)
        if k % 2 == 1:
            for col in range(1, len(headers) + 1):
                ws_sum.cell(row=ws_sum.max_row, column=col).fill = \
                    PatternFill("solid", fgColor=ALT_ROW_COLOR)

    for col in ws_sum.columns:
        sample = [c for c in col if c.value is not None]
        max_len = max((len(str(c.value)) for c in sample), default=8)
        ws_sum.column_dimensions[col[0].column_letter].width = max(max_len + 2, 14)

    if co_stark_results:
        row = ws_sum.max_row + 2
        ws_sum.cell(row=row, column=1, value="CO Stark Tuning Slopes")
        ws_sum.cell(row=row, column=1).font = Font(bold=True, size=12)
        stark_hdrs = ["Series", "Slope (cm⁻¹/V)", "Intercept (cm⁻¹)", "R²", "N Points"]
        header_row = row + 2
        for col, header in enumerate(stark_hdrs, start=1):
            ws_sum.cell(row=header_row, column=col, value=header)
        _style_header(ws_sum, header_row, len(stark_hdrs))

        for sr in co_stark_results:
            ws_sum.append([
                sr.series_name,
                round(sr.slope, 4),
                round(sr.intercept, 4),
                round(sr.r_squared, 6),
                sr.n_points,
            ])

        ws_sum.append([])
        ws_sum.append(["Series-wise data (Potential vs Center):"])
        ws_sum.cell(ws_sum.max_row, 1).font = Font(bold=True)
        for sr in co_stark_results:
            ws_sum.append([])
            ws_sum.append([sr.series_name])
            ws_sum.cell(ws_sum.max_row, 1).font = Font(bold=True)
            ws_sum.append(["Potential (V)", "Center (cm⁻¹)"])
            _style_header(ws_sum, ws_sum.max_row, 2)
            for pot, ctr in zip(sr.potentials, sr.centers):
                ws_sum.append([round(pot, 4), round(ctr, 4)])

    # ── CO manual deconvolution 시트 (per spectrum) ───────────
    for record in co_fit_records:
        fname  = record['filename']
        entry  = entry_map.get(fname)
        if entry is None:
            continue

        entry_state = co_states.get(entry.filepath, {})
        manual_state = entry_state.get('manual_fit', {})
        if not manual_state:
            manual_state = entry_state.get('CO_B', {})
        raw_fit = (
            manual_state.get('fit_result')
            or manual_state.get('raw_fit_result')
        )
        wn_b = manual_state.get('wn')
        if wn_b is None:
            wn_b = manual_state.get('wn_b')
        ab_pos_b = manual_state.get('ab')
        if ab_pos_b is None:
            ab_pos_b = manual_state.get('ab_pos_b')

        if raw_fit is None or not raw_fit.success or wn_b is None or ab_pos_b is None:
            continue
        if len(raw_fit.peaks) < 1:
            continue

        cleaned_stem = re.sub(r'[\\/*?:\[\]]', '_', Path(fname).stem)[:22]
        sheet_name = f"CO_Fit_{cleaned_stem}"
        existing = {ws.title for ws in wb.worksheets}
        base, n = sheet_name, 1
        while sheet_name in existing:
            sheet_name = f"{base[:22]}_{n}"; n += 1

        ws = wb.create_sheet(sheet_name)
        ws['A1'] = fname
        ws['A1'].font = Font(bold=True)
        pot = potentials.get(fname)
        if pot is not None:
            ws['B1'] = f"Potential: {pot:.3f} V"

        assignments = list(manual_state.get('assignments') or [])
        data_hdrs = ["Wavenumber (cm⁻¹)", "Corrected", "Fitted Total"]
        for i, peak in enumerate(raw_fit.peaks):
            label = assignments[i] if i < len(assignments) else "Unassigned"
            data_hdrs.append(f"P{i + 1} {label} ({peak.center:.0f} cm⁻¹)")
        ws.append([])
        ws.append(data_hdrs)
        _style_header(ws, ws.max_row, len(data_hdrs))

        for j in range(len(wn_b)):
            row = [
                round(wn_b[j], 4),
                round(ab_pos_b[j], 6),
                round(raw_fit.fitted_curve[j], 6),
            ]
            for curve in raw_fit.individual_curves:
                row.append(round(curve[j], 6))
            ws.append(row)

        _autosize_columns(ws, sample_row_limit=5, min_width=12)

    wb.save(filepath)
    return filepath


def export_batch(batch_results: list[dict], filepath: str):
    """
    배치 결과를 Excel로 저장.
    batch_results: [{'filename': str, 'fit_result': FitResult, ...}, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Results"

    # 동적 헤더 (최대 피크 수 기준)
    max_peaks = max(len(r['fit_result'].peaks) for r in batch_results if r['fit_result'].success)

    headers = ["File", "R²"]
    for i in range(max_peaks):
        headers += [f"P{i+1} Center (cm⁻¹)", f"P{i+1} FWHM", f"P{i+1} Area", f"P{i+1} Area%"]

    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for k, r in enumerate(batch_results):
        row = [r['filename'], "FAILED"]
        fr = r['fit_result']
        if fr.success:
            row[1] = round(fr.r_squared, 4)
            for p in fr.peaks:
                row += [round(p.center, 2), round(p.fwhm, 2),
                         round(p.area, 6), round(p.area_fraction, 2)]
        ws.append(row)
        if k % 2 == 1:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = \
                    PatternFill("solid", fgColor=ALT_ROW_COLOR)

    # 컬럼 너비 자동
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)

    wb.save(filepath)
    return filepath
