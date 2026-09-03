import os
import sys
import tempfile
import unittest
from pathlib import Path

import numpy as np

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "ir_analyzer"))

from PyQt5.QtWidgets import QApplication, QComboBox, QMessageBox, QTableWidget
from openpyxl import load_workbook

from core.exporter import export_spectra_excel
from core.peak_finder import PeakGuess
from ui.main_window import MainWindow
from ui.analysis_widget import AnalysisWidget
from ui.plot_widget import (
    AXIS_LABEL_FONT_SIZE_PX,
    AXIS_TEXT_COLOR,
    AXIS_TICK_FONT_SIZE,
    PlotWidget,
)
from ui.right_panel import RightPanel
from ui.spectrum_list import SpectrumEntry, SpectrumListWidget


class SpectrumListUiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication([])

    def test_sidebar_renders_potential_assignments_table(self):
        widget = SpectrumListWidget()

        self.assertIsNotNone(
            widget.findChild(object, "spectra_potential_table"),
            "Potential Assignments table should remain in the spectra sidebar",
        )

    def test_analysis_view_does_not_render_potential_assignment_sidebar(self):
        widget = AnalysisWidget()

        self.assertIsNone(
            widget.findChild(object, "analysis_assignments_table"),
            "Potential Assignments table should not be rendered in Analysis",
        )
        self.assertIsNone(
            widget.findChild(object, "analysis_focus_card"),
            "Focused Spectrum card should not be rendered in Analysis",
        )

    def test_spectrum_plot_axes_use_larger_white_fonts(self):
        widget = PlotWidget()
        try:
            for axis_name in ("bottom", "left"):
                axis = widget.pw.getAxis(axis_name)
                tick_font = axis.style["tickFont"]
                self.assertEqual(tick_font.pointSize(), AXIS_TICK_FONT_SIZE)
                self.assertFalse(tick_font.bold())
                self.assertEqual(
                    axis.labelStyle.get("font-size"),
                    f"{AXIS_LABEL_FONT_SIZE_PX}px",
                )
                self.assertNotIn("font-weight", axis.labelStyle)
                self.assertEqual(axis.labelStyle.get("color"), AXIS_TEXT_COLOR)
                self.assertEqual(axis.textPen().color().name(), AXIS_TEXT_COLOR)
        finally:
            widget.close()

    def test_analysis_compare_tab_exposes_peak_area_metrics(self):
        widget = AnalysisWidget()
        combo = widget.findChild(QComboBox, "analysis_compare_metric_combo")

        self.assertIsNotNone(combo)
        self.assertEqual(combo.currentData(), "P3+P4")
        self.assertEqual(
            [combo.itemData(i) for i in range(combo.count())],
            ["P1", "P2", "P3", "P4", "P1+P2", "P3+P4"],
        )

    def test_analysis_exposes_integrated_area_tab_and_table(self):
        widget = AnalysisWidget()
        try:
            table = widget.findChild(QTableWidget, "analysis_integrated_area_table")

            self.assertIsNotNone(table)
            widget.set_current_subtab("Integrated Areas")
            self.assertEqual(widget.get_current_subtab(), "Integrated Areas")

            widget.update_integrated_areas([
                {
                    "region_name": "NO3",
                    "filename": "sample.dpt",
                    "potential": -0.1,
                    "area": 1.23,
                    "wn_min": 1300.0,
                    "wn_max": 1450.0,
                    "source": "corrected",
                }
            ])

            self.assertEqual(table.rowCount(), 1)
            self.assertEqual(table.item(0, 0).text(), "NO3")
            self.assertEqual(table.item(0, 6).text(), "corrected")
        finally:
            widget.close()

    def test_workspace_save_payload_reopens_as_separate_session_tabs(self):
        source = MainWindow()
        target = MainWindow()
        try:
            wn = np.linspace(1000.0, 1400.0, 41)
            entry_a = SpectrumEntry(
                "/tmp/session-a.dpt",
                "Alpha :: session-a.dpt",
                wn,
                np.ones_like(wn),
                "#89b4fa",
                original_name="session-a.dpt",
                source_session_label="Alpha",
                source_spectrum_path="/tmp/session-a.dpt",
            )
            entry_b = SpectrumEntry(
                "/tmp/session-b.dpt",
                "Beta :: session-b.dpt",
                wn,
                np.ones_like(wn) * 2,
                "#fab387",
                original_name="session-b.dpt",
                source_session_label="Beta",
                source_spectrum_path="/tmp/session-b.dpt",
            )
            source.spectrum_list.add_entry(entry_a, select=False, emit_signal=False)
            source.spectrum_list.add_entry(entry_b, select=False, emit_signal=False)
            source._total_shifts = {
                "Alpha": {entry_a.name: 0.1},
                "Beta": {entry_b.name: 0.2},
            }
            source._total_inactive_ranges = {
                "Alpha": [(1100.0, 1120.0)],
                "Beta": [(1200.0, 1220.0)],
            }

            payload = source._build_session_payload([entry_a, entry_b])
            imported = target._merge_session_data(payload, "/tmp/workspace.irsession")

            self.assertEqual(imported, 2)
            self.assertIn("Alpha", target.spectrum_list.get_session_keys())
            self.assertIn("Beta", target.spectrum_list.get_session_keys())
            self.assertNotIn("workspace", target.spectrum_list.get_session_keys())
            imported_by_name = {
                entry.original_name: entry
                for entry in target.spectrum_list.get_all_entries()
            }
            self.assertEqual(
                target.spectrum_list.get_session_key_for_entry(
                    imported_by_name["session-a.dpt"]),
                "Alpha",
            )
            self.assertEqual(
                target.spectrum_list.get_session_key_for_entry(
                    imported_by_name["session-b.dpt"]),
                "Beta",
            )
            self.assertEqual(target._total_shifts["Alpha"]["Alpha :: session-a.dpt"], 0.1)
            self.assertEqual(target._total_shifts["Beta"]["Beta :: session-b.dpt"], 0.2)
            self.assertEqual(target._total_inactive_ranges["Alpha"], [(1100.0, 1120.0)])
            self.assertEqual(target._total_inactive_ranges["Beta"], [(1200.0, 1220.0)])
        finally:
            source.close()
            target.close()

    def test_single_session_import_still_uses_file_stem_as_session_tab(self):
        source = MainWindow()
        target = MainWindow()
        try:
            wn = np.linspace(1000.0, 1400.0, 41)
            entry = SpectrumEntry(
                "/tmp/original-session.dpt",
                "Original Tab :: original-session.dpt",
                wn,
                np.ones_like(wn),
                "#89b4fa",
                original_name="original-session.dpt",
                source_session_label="Original Tab",
                source_spectrum_path="/tmp/original-session.dpt",
            )
            source.spectrum_list.add_entry(entry, select=False, emit_signal=False)

            payload = source._build_session_payload([entry])
            target._merge_session_data(payload, "/tmp/Renamed Session.irsession")

            session_keys = target.spectrum_list.get_session_keys()
            self.assertIn("Renamed Session", session_keys)
            self.assertNotIn("Original Tab", session_keys)
            imported = target.spectrum_list.get_all_entries()[0]
            self.assertEqual(
                target.spectrum_list.get_session_key_for_entry(imported),
                "Renamed Session",
            )
            self.assertEqual(imported.name, "Renamed Session :: original-session.dpt")
        finally:
            source.close()
            target.close()

    def test_same_dpt_file_can_be_loaded_into_different_session_tabs(self):
        widget = SpectrumListWidget()
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                dpt_path = Path(tmpdir) / "shared.dpt"
                dpt_path.write_text(
                    "1400 0.1\n1300 0.2\n1200 0.3\n",
                    encoding="utf-8",
                )

                first = widget.add_file(str(dpt_path))
                session_key = widget.create_workspace()
                second = widget.add_file(str(dpt_path))

                self.assertIsNotNone(first)
                self.assertIsNotNone(second)
                self.assertEqual(len(widget.get_all_entries()), 2)
                self.assertEqual(
                    widget.get_session_key_for_entry(first),
                    widget.LOOSE_FILES_KEY,
                )
                self.assertEqual(
                    widget.get_session_key_for_entry(second),
                    session_key,
                )
                self.assertEqual(first.source_spectrum_path, str(dpt_path))
                self.assertEqual(second.source_spectrum_path, str(dpt_path))
                self.assertNotEqual(first.filepath, second.filepath)
        finally:
            widget.close()

    def test_same_dpt_file_is_still_ignored_within_one_session_tab(self):
        widget = SpectrumListWidget()
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                dpt_path = Path(tmpdir) / "duplicate.dpt"
                dpt_path.write_text(
                    "1400 0.1\n1300 0.2\n1200 0.3\n",
                    encoding="utf-8",
                )

                first = widget.add_file(str(dpt_path))
                second = widget.add_file(str(dpt_path))

                self.assertIsNotNone(first)
                self.assertIsNone(second)
                self.assertEqual(len(widget.get_all_entries()), 1)
        finally:
            widget.close()

    def test_analysis_subtab_is_preserved_when_returning_from_spectrum(self):
        window = MainWindow()
        try:
            window.analysis_widget.set_current_subtab("Compare")
            window.center_tabs.setCurrentWidget(window.plot_widget)
            window.center_tabs.setCurrentWidget(window.analysis_widget)

            self.assertEqual(window.analysis_widget.get_current_subtab(), "Compare")
        finally:
            window.close()

    def test_baseline_controls_are_total_only(self):
        window = MainWindow()
        try:
            self.assertTrue(window.right_panel._baseline_group.isHidden())
            self.assertEqual(window.right_panel.combo_bl_algo.currentText(), "Manual")

            window.right_panel.set_mode("Total")
            self.assertFalse(window.right_panel._baseline_group.isHidden())

            window.right_panel.set_mode("CO")
            self.assertTrue(window.right_panel._baseline_group.isHidden())
            self.assertFalse(window.plot_widget.get_co_endpoints())
        finally:
            window.close()

    def test_co_mode_hides_user_analysis_region_controls(self):
        window = MainWindow()
        try:
            window.right_panel.set_mode("CO")

            self.assertTrue(window.right_panel._region_group.isHidden())

            window.right_panel.set_mode("Total")

            self.assertFalse(window.right_panel._region_group.isHidden())
        finally:
            window.close()

    def test_co_peak_table_preserves_assignment_and_locks(self):
        panel = RightPanel()
        try:
            guess = PeakGuess(center=1850.0, amplitude=0.004, sigma=45.0, index=0)
            guess.assignment = "CO_B"

            panel.set_co_guesses(
                [guess],
                locks=[{"center": True, "amplitude": False, "sigma": True}],
            )

            restored = panel.get_co_guesses()
            locks = panel.get_co_locks()

            self.assertEqual(len(restored), 1)
            self.assertEqual(restored[0].assignment, "CO_B")
            self.assertEqual(restored[0].shape, "gaussian")
            self.assertTrue(locks[0]["center"])
            self.assertFalse(locks[0]["amplitude"])
            self.assertTrue(locks[0]["sigma"])
        finally:
            panel.close()

    def test_co_manual_peak_created_adds_guess_to_table_and_state(self):
        window = MainWindow()
        try:
            wn = np.linspace(1400.0, 2230.0, 240)
            ab = 0.012 * np.exp(-((wn - 1655.0) / 28.0) ** 2)
            entry = SpectrumEntry(
                "/tmp/co-manual-add.dpt",
                "co-manual-add.dpt",
                wn,
                ab,
                "#89b4fa",
            )
            window._current_entry = entry
            window.right_panel.set_mode("CO")

            window._on_co_peak_created(1660.0, 0.01, 32.0)

            guesses = window.right_panel.get_co_guesses()
            manual = window._co_states[entry.filepath]["manual_fit"]
            self.assertEqual(len(guesses), 1)
            self.assertAlmostEqual(guesses[0].center, 1660.0)
            self.assertEqual(guesses[0].assignment, "Unassigned")
            self.assertEqual(len(manual["guesses"]), 1)
            self.assertEqual(manual["assignments"], ["Unassigned"])
        finally:
            window.close()

    def test_clear_co_peaks_removes_manual_legacy_and_records(self):
        window = MainWindow()
        try:
            wn = np.linspace(1400.0, 2230.0, 240)
            entry = SpectrumEntry(
                "/tmp/co-clear.dpt",
                "co-clear.dpt",
                wn,
                np.zeros_like(wn),
                "#89b4fa",
            )
            guess = PeakGuess(center=1660.0, amplitude=0.01, sigma=30.0, index=0)
            window._current_entry = entry
            window.right_panel.set_mode("CO")
            window._co_states[entry.filepath] = {
                "manual_fit": {
                    "guesses": [guess],
                    "assignments": ["CO_B"],
                    "locks": [{"center": False, "amplitude": False, "sigma": False}],
                },
                "CO_B": {
                    "guesses": [guess],
                    "fit_result": object(),
                },
            }
            window._co_fit_records = [{"filename": entry.name, "CO_L": None, "CO_B": object()}]

            window._on_co_peaks_cleared()

            state = window._co_states[entry.filepath]
            self.assertNotIn("guesses", state["manual_fit"])
            self.assertNotIn("guesses", state["CO_B"])
            self.assertEqual(window.right_panel.co_init_table.rowCount(), 0)
            self.assertEqual(window._co_fit_records, [])
        finally:
            window.close()

    def test_saved_co_deconvolution_does_not_force_always_two_peak_mode(self):
        window = MainWindow()
        try:
            wn = np.linspace(1400.0, 2230.0, 101)
            entry = SpectrumEntry(
                "/tmp/co-deconv-mode.dpt",
                "co-deconv-mode.dpt",
                wn,
                np.exp(-((wn - 1660.0) / 30.0) ** 2),
                "#89b4fa",
            )
            window.spectrum_list.add_entry(entry)
            window._current_entry = entry
            window._co_states[entry.filepath] = {
                "CO_B": {
                    "analysis_mode": "deconv",
                    "guesses": [object()],
                }
            }

            window._sync_co_b_mode_from_saved_state()

            self.assertEqual(window.right_panel.get_co_b_fit_mode(), "auto")
        finally:
            window.close()

    def test_co_view_uses_full_spectrum_region_not_user_region_inputs(self):
        window = MainWindow()
        try:
            calls = []
            original_analysis = window._analysis_arrays_for_entry
            original_zoom = window.plot_widget.zoom_to

            def fake_analysis(entry, wn_min, wn_max):
                calls.append(("analysis", wn_min, wn_max))
                wn = np.linspace(float(wn_min), float(wn_max), 5)
                return wn, wn * 0.0, wn * 0.0, wn * 0.0, "raw"

            def fake_zoom(wn_min, wn_max, *args, **kwargs):
                calls.append(("zoom", wn_min, wn_max))

            window._analysis_arrays_for_entry = fake_analysis
            window.plot_widget.zoom_to = fake_zoom
            wn = np.linspace(1000.0, 3990.0, 101)
            entry = SpectrumEntry(
                "/tmp/co-fixed-view.dpt", "co-fixed-view.dpt", wn, wn * 0.0, "#89b4fa")
            window.right_panel.set_mode("CO")
            window.right_panel.spin_wn_min.setValue(1000.0)
            window.right_panel.spin_wn_max.setValue(3990.0)

            window._apply_co_view(entry)

            self.assertIn(("analysis", 1000.0, 3990.0), calls)
            self.assertIn(("zoom", 1000.0, 3990.0), calls)
            window._analysis_arrays_for_entry = original_analysis
            window.plot_widget.zoom_to = original_zoom
        finally:
            window.close()

    def test_co_b_refinement_data_uses_full_spectrum_region(self):
        window = MainWindow()
        try:
            calls = []
            original_analysis = window._analysis_arrays_for_entry

            def fake_analysis(entry, wn_min, wn_max):
                calls.append((wn_min, wn_max))
                wn = np.linspace(float(wn_min), float(wn_max), 5)
                return wn, wn * 0.0, wn * 0.0, wn * 0.0, "raw"

            window._analysis_arrays_for_entry = fake_analysis
            wn = np.linspace(900.0, 3990.0, 101)
            entry = SpectrumEntry(
                "/tmp/co-b-full-fit.dpt", "co-b-full-fit.dpt", wn, wn * 0.0, "#89b4fa")

            window._co_b_fit_data(entry)

            self.assertEqual(calls[-1], (900.0, 3990.0))
            window._analysis_arrays_for_entry = original_analysis
        finally:
            window.close()

    def test_co_reanalyze_auto_without_guesses_does_not_invent_global_co_b_peak(self):
        window = MainWindow()
        try:
            wn = np.linspace(1000.0, 3990.0, 400)
            ab = (
                0.02 * np.exp(-((wn - 3480.0) / 80.0) ** 2)
                + 0.005 * np.exp(-((wn - 1660.0) / 25.0) ** 2)
            )
            entry = SpectrumEntry(
                "/tmp/co-auto-no-guesses.dpt",
                "co-auto-no-guesses.dpt",
                wn,
                ab,
                "#89b4fa",
            )
            window.right_panel.set_mode("CO")
            window.right_panel.set_co_b_fit_mode("auto")

            outcome = window._fit_co_entry(entry)

            self.assertIsNone(outcome["co_b"])
            co_b_state = window._co_states[entry.filepath]["CO_B"]
            self.assertIsNone(co_b_state["fit_result"])
            self.assertEqual(co_b_state["status"], "review")
        finally:
            window.close()

    def test_co_fit_uses_explicit_assignment_for_co_b(self):
        window = MainWindow()
        try:
            wn = np.linspace(1400.0, 2230.0, 240)
            ab = (
                0.012 * np.exp(-((wn - 1655.0) / 28.0) ** 2)
                + 0.004 * np.exp(-((wn - 1850.0) / 45.0) ** 2)
            )
            entry = SpectrumEntry(
                "/tmp/co-manual-guesses.dpt",
                "co-manual-guesses.dpt",
                wn,
                ab,
                "#89b4fa",
            )
            window.right_panel.set_mode("CO")
            p1 = PeakGuess(center=1655.0, amplitude=0.012, sigma=28.0, index=0)
            p1.assignment = "OH bending"
            p2 = PeakGuess(center=1850.0, amplitude=0.004, sigma=45.0, index=1)
            p2.assignment = "CO_B"
            window._co_states[entry.filepath] = {
                "manual_fit": {
                    "guesses": [
                        p1,
                        p2,
                    ],
                    "assignments": ["OH bending", "CO_B"],
                    "locks": [
                        {"center": False, "amplitude": False, "sigma": False},
                        {"center": False, "amplitude": False, "sigma": False},
                    ],
                }
            }

            outcome = window._fit_co_entry(entry)

            self.assertTrue(outcome["used_deconv"])
            self.assertIsNotNone(outcome["co_b"])
            self.assertTrue(outcome["co_b"].success)
            self.assertGreater(outcome["co_b"].peaks[0].center, 1700.0)
            self.assertIsNone(outcome["co_l"])
            self.assertIn("fit_result", window._co_states[entry.filepath]["manual_fit"])
        finally:
            window.close()

    def test_unedited_auto_baseline_states_are_discarded_on_load(self):
        window = MainWindow()
        try:
            states = {
                "auto": {
                    "algo": "Auto Baseline",
                    "manual_override": False,
                    "points": [(1500.0, 0.1), (4000.0, 0.2)],
                },
                "manual": {
                    "algo": "Manual",
                    "manual_override": True,
                    "points": [(1500.0, 0.1), (4000.0, 0.2)],
                },
            }

            sanitized = window._sanitize_total_baseline_states(states)

            self.assertNotIn("auto", sanitized)
            self.assertIn("manual", sanitized)
        finally:
            window.close()

    def test_analysis_prefers_total_corrected_then_falls_back_to_raw(self):
        window = MainWindow()
        try:
            wn = np.linspace(1500.0, 4000.0, 251)
            raw = 0.01 + wn * 1e-6
            corrected = raw - 0.004
            entry = SpectrumEntry(
                "/tmp/analysis-source.dpt",
                "analysis-source.dpt",
                wn,
                raw,
                "#89b4fa",
            )
            window._total_baseline_states[entry.filepath] = {
                "wn": wn.copy(),
                "raw": raw.copy(),
                "baseline": np.full_like(raw, 0.004),
                "corrected": corrected.copy(),
            }

            out_wn, out_raw, baseline, out_corrected, source = (
                window._analysis_arrays_for_entry(entry, 3000.0, 3990.0)
            )
            mask = (wn >= 3000.0) & (wn <= 3990.0)
            self.assertEqual(source, "total")
            np.testing.assert_allclose(out_wn, wn[mask])
            np.testing.assert_allclose(out_raw, raw[mask])
            np.testing.assert_allclose(baseline, 0.004)
            np.testing.assert_allclose(out_corrected, corrected[mask])

            window._total_baseline_states.clear()
            _, out_raw, baseline, out_corrected, source = (
                window._analysis_arrays_for_entry(entry, 3000.0, 3990.0)
            )
            self.assertEqual(source, "raw")
            np.testing.assert_allclose(baseline, 0.0)
            np.testing.assert_allclose(out_corrected, out_raw)
        finally:
            window.close()

    def test_total_integral_uses_corrected_data_and_updates_analysis(self):
        window = MainWindow()
        try:
            wn = np.array([1200.0, 1300.0, 1400.0, 1450.0, 1500.0])
            raw_one = np.full_like(wn, 5.0)
            raw_two = np.full_like(wn, 4.0)
            corrected_one = np.full_like(wn, 2.0)
            corrected_two = np.full_like(wn, 1.0)
            entry_one = SpectrumEntry(
                "/tmp/integral-one.dpt", "integral-one.dpt", wn, raw_one, "#89b4fa")
            entry_two = SpectrumEntry(
                "/tmp/integral-two.dpt", "integral-two.dpt", wn, raw_two, "#fab387")
            window.spectrum_list.add_entry(entry_one, select=False, emit_signal=False)
            window.spectrum_list.add_entry(entry_two, select=False, emit_signal=False)
            window._current_entry = entry_one
            window.spectrum_list.set_potentials(
                [entry_one.name, entry_two.name],
                {entry_one.name: -0.1, entry_two.name: -0.2},
                emit_changed=False,
            )
            window._total_baseline_states = {
                entry_one.filepath: {
                    "wn": wn.copy(),
                    "raw": raw_one.copy(),
                    "baseline": raw_one - corrected_one,
                    "corrected": corrected_one.copy(),
                    "points": [],
                    "algo": "Manual",
                    "params": {},
                    "region": (1200.0, 1500.0),
                    "manual_override": True,
                },
                entry_two.filepath: {
                    "wn": wn.copy(),
                    "raw": raw_two.copy(),
                    "baseline": raw_two - corrected_two,
                    "corrected": corrected_two.copy(),
                    "points": [],
                    "algo": "Manual",
                    "params": {},
                    "region": (1200.0, 1500.0),
                    "manual_override": True,
                },
            }
            window.right_panel.set_mode("Total")
            window.right_panel.set_total_integral_config({
                "name": "NO3",
                "wn_min": 1300.0,
                "wn_max": 1450.0,
            })

            window._calculate_total_integral()

            session_key = window.spectrum_list.get_current_session_filter()
            rows = window._total_integral_results[session_key]["NO3"]
            self.assertEqual(len(rows), 2)
            self.assertEqual({row["source"] for row in rows}, {"corrected"})
            areas = {row["filename"]: row["area"] for row in rows}
            self.assertAlmostEqual(areas[entry_one.name], 300.0)
            self.assertAlmostEqual(areas[entry_two.name], 150.0)
            self.assertEqual(window.analysis_widget.integral_table.rowCount(), 2)
            self.assertEqual(window.analysis_widget.get_current_subtab(), "Integrated Areas")
        finally:
            window.close()

    def test_total_integral_falls_back_to_raw_without_total_baseline(self):
        window = MainWindow()
        try:
            wn = np.array([1200.0, 1300.0, 1400.0, 1450.0, 1500.0])
            raw = np.full_like(wn, 3.0)
            entry = SpectrumEntry(
                "/tmp/integral-raw.dpt", "integral-raw.dpt", wn, raw, "#89b4fa")
            window.spectrum_list.add_entry(entry, select=False, emit_signal=False)
            window._current_entry = entry
            window.right_panel.set_mode("Total")
            window.right_panel.set_total_integral_config({
                "name": "NO3",
                "wn_min": 1300.0,
                "wn_max": 1450.0,
            })

            window._calculate_total_integral()

            session_key = window.spectrum_list.get_current_session_filter()
            row = window._total_integral_results[session_key]["NO3"][0]
            self.assertEqual(row["source"], "raw")
            self.assertAlmostEqual(row["area"], 450.0)
        finally:
            window.close()

    def test_total_integral_state_is_included_in_session_payload(self):
        window = MainWindow()
        try:
            wn = np.array([1200.0, 1300.0, 1400.0, 1450.0, 1500.0])
            raw = np.full_like(wn, 3.0)
            entry = SpectrumEntry(
                "/tmp/integral-payload.dpt", "integral-payload.dpt", wn, raw, "#89b4fa")
            window.spectrum_list.add_entry(entry, select=False, emit_signal=False)
            window._current_entry = entry
            window.right_panel.set_mode("Total")
            window.right_panel.set_total_integral_config({
                "name": "NO3",
                "wn_min": 1300.0,
                "wn_max": 1450.0,
            })
            window._calculate_total_integral()

            payload = window._build_session_payload([entry])
            session_key = window.spectrum_list.get_current_session_filter()

            self.assertIn(session_key, payload["total_integral_regions"])
            self.assertIn("NO3", payload["total_integral_regions"][session_key])
            self.assertIn(session_key, payload["total_integral_results"])
            self.assertEqual(
                payload["total_integral_results"][session_key]["NO3"][0]["filename"],
                entry.name,
            )
        finally:
            window.close()

    def test_total_manual_baseline_updates_corrected_curve_while_editing(self):
        window = MainWindow()
        try:
            wn = np.linspace(1500.0, 4000.0, 251)
            raw = 0.01 + 0.02 * np.exp(-((wn - 3350.0) / 200.0) ** 2)
            entry = SpectrumEntry(
                "/tmp/live-baseline.dpt", "live-baseline.dpt", wn, raw, "#89b4fa")
            window._current_entry = entry
            window.right_panel.set_wavenumber_range(1500.0, 4000.0)
            window.right_panel.set_mode("Total")
            window.right_panel.combo_bl_algo.setCurrentText("Manual")
            window.right_panel.btn_edit_bl.setChecked(True)

            window._on_baseline_point_added(1600.0, 999.0)
            window._on_baseline_point_added(3900.0, -999.0)

            state = window._total_baseline_states[entry.filepath]
            self.assertFalse(np.allclose(state["corrected"], state["raw"]))
            self.assertIn("raw", window.plot_widget._items)
            self.assertIn("baseline", window.plot_widget._items)
            active_spec = window.plot_widget._total_specs[0]
            np.testing.assert_allclose(active_spec["ab"], state["corrected"])
        finally:
            window.close()

    def test_total_overlay_aligns_unprocessed_spectra_with_different_offsets(self):
        window = MainWindow()
        try:
            wn = np.linspace(1500.0, 4000.0, 251)
            peak = 0.02 * np.exp(-((wn - 3350.0) / 200.0) ** 2)
            entry_one = SpectrumEntry(
                "/tmp/offset-one.dpt", "offset-one.dpt", wn, 0.01 + peak, "#89b4fa")
            entry_two = SpectrumEntry(
                "/tmp/offset-two.dpt", "offset-two.dpt", wn, 0.04 + peak, "#fab387")

            display_one = window._total_comparison_display(
                entry_one, wn, entry_one.absorbance)
            display_two = window._total_comparison_display(
                entry_two, wn, entry_two.absorbance)

            np.testing.assert_allclose(display_one, display_two, atol=1e-12)
        finally:
            window.close()

    def test_total_overlay_keeps_effective_corrected_spectrum_unchanged(self):
        window = MainWindow()
        try:
            wn = np.linspace(1500.0, 4000.0, 251)
            corrected = 0.02 * np.exp(-((wn - 3350.0) / 200.0) ** 2)
            entry = SpectrumEntry(
                "/tmp/corrected-overlay.dpt",
                "corrected-overlay.dpt",
                wn,
                corrected + 0.03,
                "#89b4fa",
            )
            window._total_baseline_states[entry.filepath] = {
                "wn": wn.copy(),
                "raw": entry.absorbance.copy(),
                "baseline": np.full_like(wn, 0.03),
                "corrected": corrected.copy(),
                "points": [(1500.0, 0.03), (4000.0, 0.03)],
                "algo": "Manual",
                "manual_override": True,
            }

            displayed = window._total_comparison_display(entry, wn, corrected)

            np.testing.assert_allclose(displayed, corrected)
        finally:
            window.close()

    def test_total_normalize_preserves_zero_and_scales_corrected_spectra_for_overlay(self):
        window = MainWindow()
        try:
            window.right_panel.set_mode("Total")
            window.right_panel.combo_oh_overlay_intensity.setCurrentText("Normalize")
            window.right_panel.spin_oh_norm_min.setValue(3000.0)
            window.right_panel.spin_oh_norm_max.setValue(3990.0)
            wn = np.linspace(1500.0, 4000.0, 251)
            shape = np.exp(-((wn - 3350.0) / 200.0) ** 2)
            corrected_one = 0.20 * shape
            corrected_two = 0.05 * shape
            entry_one = SpectrumEntry(
                "/tmp/norm-one.dpt", "norm-one.dpt", wn, corrected_one, "#89b4fa")
            entry_two = SpectrumEntry(
                "/tmp/norm-two.dpt", "norm-two.dpt", wn, corrected_two, "#fab387")
            for entry, corrected in ((entry_one, corrected_one), (entry_two, corrected_two)):
                window._total_baseline_states[entry.filepath] = {
                    "wn": wn.copy(),
                    "raw": corrected.copy(),
                    "baseline": np.full_like(wn, 0.01),
                    "corrected": corrected.copy(),
                    "points": [(1500.0, 0.01), (4000.0, 0.01)],
                    "algo": "Manual",
                    "manual_override": True,
                }

            display_one = window._total_comparison_display(
                entry_one, wn, corrected_one)
            display_two = window._total_comparison_display(
                entry_two, wn, corrected_two)

            mask = (wn >= 3000.0) & (wn <= 3990.0)
            self.assertAlmostEqual(float(np.nanmax(display_one[mask])), 1.0, places=6)
            self.assertAlmostEqual(float(np.nanmax(display_two[mask])), 1.0, places=6)
            self.assertAlmostEqual(float(display_one[0]), 0.0, places=5)
            self.assertAlmostEqual(float(display_two[0]), 0.0, places=5)
            np.testing.assert_allclose(display_one, display_two, atol=1e-6)
        finally:
            window.close()

    def test_total_normalize_uses_saved_shift_only_when_shift_mode_is_on(self):
        window = MainWindow()
        try:
            wn = np.linspace(1500.0, 4000.0, 251)
            entry = SpectrumEntry(
                "/tmp/shifted-normalize.dpt", "shifted-normalize.dpt",
                wn, np.exp(-((wn - 3350.0) / 200.0) ** 2), "#89b4fa")
            window.spectrum_list.add_entry(entry)
            window.spectrum_list.list_widget.item(0).setSelected(True)
            window.right_panel.set_mode("Total")
            window.right_panel.combo_oh_overlay_intensity.setCurrentText("Normalize")
            session_key = window.spectrum_list.get_current_session_filter()
            window._total_shifts[session_key] = {entry.name: 0.25}

            self.assertTrue(window.right_panel.cb_total_shift.isEnabled())
            self.assertEqual(window._build_total_specs()[0]["shift"], 0.0)

            window.right_panel.cb_total_shift.setChecked(True)

            self.assertEqual(window._build_total_specs()[0]["shift"], 0.25)
        finally:
            window.close()

    def test_total_original_uses_saved_shift_only_when_shift_mode_is_on(self):
        window = MainWindow()
        try:
            wn = np.linspace(1500.0, 4000.0, 251)
            entry = SpectrumEntry(
                "/tmp/original-shift.dpt", "original-shift.dpt",
                wn, np.exp(-((wn - 3350.0) / 200.0) ** 2), "#89b4fa")
            window.spectrum_list.add_entry(entry)
            window.spectrum_list.list_widget.item(0).setSelected(True)
            window.right_panel.set_mode("Total")
            window.right_panel.combo_oh_overlay_intensity.setCurrentText("Original")
            session_key = window.spectrum_list.get_current_session_filter()
            window._total_shifts[session_key] = {entry.name: 0.25}

            self.assertEqual(window._build_total_specs()[0]["shift"], 0.0)

            window.right_panel.cb_total_shift.setChecked(True)

            self.assertEqual(window._build_total_specs()[0]["shift"], 0.25)
        finally:
            window.close()

    def test_total_baseline_points_are_visible_while_editing_even_if_baseline_curve_hidden(self):
        window = MainWindow()
        try:
            wn = np.linspace(1500.0, 4000.0, 251)
            raw = 0.01 + 0.02 * np.exp(-((wn - 3350.0) / 200.0) ** 2)
            entry = SpectrumEntry(
                "/tmp/visible-points.dpt", "visible-points.dpt", wn, raw, "#89b4fa")
            window.spectrum_list.add_entry(entry)
            window._current_entry = entry
            window.right_panel.set_mode("Total")
            window.plot_widget.cb_baseline.setChecked(False)
            window._total_baseline_states[entry.filepath] = {
                "wn": wn.copy(),
                "raw": raw.copy(),
                "baseline": np.zeros_like(wn),
                "corrected": raw.copy(),
                "points": [(1600.0, 0.01), (3900.0, 0.02)],
                "algo": "Manual",
                "manual_override": True,
            }
            window.plot_widget.set_baseline_edit_mode(True)

            window._restore_total_baseline_points_for_current()

            self.assertEqual(len(window.plot_widget._bl_scatter_items), 2)
            self.assertTrue(all(item.isVisible() for item in window.plot_widget._bl_scatter_items))
            restored_y = [
                float(item.getData()[1][0])
                for item in window.plot_widget._bl_scatter_items
            ]
            np.testing.assert_allclose(restored_y, [0.01, 0.02])
        finally:
            window.close()

    def test_total_baseline_point_edit_preserves_zoomed_view(self):
        window = MainWindow()
        try:
            wn = np.linspace(1500.0, 4000.0, 251)
            raw = 0.01 + 0.02 * np.exp(-((wn - 3350.0) / 200.0) ** 2)
            entry = SpectrumEntry(
                "/tmp/preserve-view.dpt", "preserve-view.dpt", wn, raw, "#89b4fa")
            window._current_entry = entry
            window.right_panel.set_wavenumber_range(1500.0, 4000.0)
            window.right_panel.set_mode("Total")
            window.right_panel.btn_edit_bl.setChecked(True)
            window._on_baseline_point_added(3100.0, 999.0)
            window._on_baseline_point_added(3700.0, -999.0)

            window.plot_widget.pw.setXRange(3200.0, 3600.0, padding=0)
            window.plot_widget.pw.setYRange(-0.005, 0.025, padding=0)
            before = window.plot_widget.get_view_state()

            window._on_baseline_point_added(3400.0, 999.0)
            after = window.plot_widget.get_view_state()

            np.testing.assert_allclose(after["x_range"], before["x_range"])
            np.testing.assert_allclose(after["y_range"], before["y_range"])
        finally:
            window.close()

    def test_total_baseline_free_point_mode_preserves_clicked_y(self):
        window = MainWindow()
        try:
            wn = np.linspace(1500.0, 4000.0, 251)
            raw = np.linspace(0.01, 0.02, len(wn))
            entry = SpectrumEntry(
                "/tmp/free-baseline.dpt", "free-baseline.dpt", wn, raw, "#89b4fa")
            window._current_entry = entry
            window.right_panel.set_wavenumber_range(1500.0, 4000.0)
            window.right_panel.set_mode("Total")
            window.right_panel.btn_edit_bl.setChecked(True)
            window.right_panel.combo_bl_point_mode.setCurrentText("Free")

            window._on_baseline_point_added(3100.5, 0.1234)

            points = window._total_baseline_states[entry.filepath]["points"]
            self.assertEqual(len(points), 1)
            self.assertAlmostEqual(points[0][0], 3100.5)
            self.assertAlmostEqual(points[0][1], 0.1234)
        finally:
            window.close()

    def test_total_baseline_follow_point_mode_snaps_to_spectrum(self):
        window = MainWindow()
        try:
            wn = np.linspace(1500.0, 4000.0, 251)
            raw = np.linspace(0.01, 0.02, len(wn))
            entry = SpectrumEntry(
                "/tmp/follow-baseline.dpt", "follow-baseline.dpt", wn, raw, "#89b4fa")
            window._current_entry = entry
            window.right_panel.set_wavenumber_range(1500.0, 4000.0)
            window.right_panel.set_mode("Total")
            window.right_panel.btn_edit_bl.setChecked(True)
            window.right_panel.combo_bl_point_mode.setCurrentText("Follow Spectrum")

            window._on_baseline_point_added(3100.5, 0.1234)

            nearest = int(np.argmin(np.abs(wn - 3100.5)))
            points = window._total_baseline_states[entry.filepath]["points"]
            self.assertEqual(len(points), 1)
            self.assertAlmostEqual(points[0][0], float(wn[nearest]))
            self.assertAlmostEqual(points[0][1], float(raw[nearest]))
        finally:
            window.close()

    def test_total_baseline_survives_adding_new_workspace_and_dpt(self):
        window = MainWindow()
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                first_path = Path(tmpdir) / "first.dpt"
                second_path = Path(tmpdir) / "second.dpt"
                first_path.write_text(
                    "4000 0.010\n3600 0.030\n3200 0.012\n",
                    encoding="utf-8",
                )
                second_path.write_text(
                    "4000 0.020\n3600 0.040\n3200 0.018\n",
                    encoding="utf-8",
                )

                first = window.spectrum_list.add_file(str(first_path))
                self.assertIsNotNone(first)
                window.right_panel.set_mode("Total")
                window._current_entry = first
                window._on_baseline_point_added(4000.0, 0.010)
                window._on_baseline_point_added(3200.0, 0.012)

                saved = window._total_baseline_states[first.filepath]
                self.assertEqual(saved["algo"], "Manual")
                self.assertEqual(saved["points"], [(4000.0, 0.010), (3200.0, 0.012)])

                window.spectrum_list.create_workspace()
                second = window.spectrum_list.add_file(str(second_path))

                self.assertIsNotNone(second)
                self.assertIn(first.filepath, window._total_baseline_states)
                self.assertEqual(
                    window._total_baseline_states[first.filepath]["points"],
                    [(4000.0, 0.010), (3200.0, 0.012)],
                )
        finally:
            window.close()

    def test_auto_detect_updates_existing_oh_state_guesses(self):
        window = MainWindow()
        try:
            wn = np.linspace(3000.0, 4000.0, 401)
            raw = (
                0.015 * np.exp(-((wn - 3500.0) / 60.0) ** 2)
                + 0.007 * np.exp(-((wn - 3650.0) / 45.0) ** 2)
            )
            entry = SpectrumEntry(
                "/tmp/auto-detect-state.dpt",
                "auto-detect-state.dpt",
                wn,
                raw,
                "#89b4fa",
            )
            stale_guess = PeakGuess(center=3100.0, amplitude=0.001, sigma=30.0, index=0)
            window._current_entry = entry
            window.right_panel.set_mode("OH")
            window.right_panel.set_wavenumber_range(3000.0, 4000.0)
            window.right_panel.spin_n_peaks.setValue(2)
            window._spectrum_states[entry.filepath] = {
                "wn_crop": wn.copy(),
                "ab_crop": raw.copy(),
                "baseline": np.zeros_like(raw),
                "ab_corrected": raw.copy(),
                "fit_result": None,
                "guesses": [stale_guess],
                "locks": [],
                "baseline_points": [],
                "snapshots": [],
            }

            window._auto_detect()

            table_centers = [round(g.center, 3) for g in window.right_panel.get_guesses()]
            stored_centers = [
                round(g.center, 3)
                for g in window._spectrum_states[entry.filepath]["guesses"]
            ]
            self.assertEqual(stored_centers, table_centers)
            self.assertNotEqual(stored_centers, [3100.0])
            self.assertEqual(len(stored_centers), 2)
        finally:
            window.close()

    def test_save_current_spectrum_state_uses_current_peak_table(self):
        window = MainWindow()
        try:
            wn = np.linspace(3000.0, 4000.0, 101)
            raw = np.exp(-((wn - 3500.0) / 70.0) ** 2)
            entry = SpectrumEntry(
                "/tmp/save-current-guesses.dpt",
                "save-current-guesses.dpt",
                wn,
                raw,
                "#89b4fa",
            )
            current_guess = PeakGuess(center=3500.0, amplitude=1.0, sigma=70.0, index=0)
            stale_guess = PeakGuess(center=3100.0, amplitude=0.1, sigma=30.0, index=0)
            window._current_entry = entry
            window._wn_crop = wn.copy()
            window._ab_crop = raw.copy()
            window._baseline = np.zeros_like(raw)
            window._ab_corrected = raw.copy()
            window._baseline_points = []
            window.right_panel.set_guesses([current_guess], locks=[
                {"center": False, "amplitude": False, "sigma": False}
            ])
            window._spectrum_states[entry.filepath] = {
                "wn_crop": wn.copy(),
                "ab_crop": raw.copy(),
                "baseline": np.zeros_like(raw),
                "ab_corrected": raw.copy(),
                "fit_result": None,
                "guesses": [stale_guess],
                "locks": [],
                "baseline_points": [],
                "snapshots": [],
            }

            window._save_current_spectrum_state()

            stored = window._spectrum_states[entry.filepath]
            self.assertEqual(len(stored["guesses"]), 1)
            self.assertAlmostEqual(stored["guesses"][0].center, 3500.0)
        finally:
            window.close()

    def test_auto_fit_requires_endpoint_fit_results(self):
        window = MainWindow()
        warnings = []
        original_warning = QMessageBox.warning
        try:
            wn = np.linspace(3000.0, 4000.0, 101)
            raw = np.exp(-((wn - 3500.0) / 70.0) ** 2)
            entries = [
                SpectrumEntry(f"/tmp/auto-fit-{i}.dpt", f"auto-fit-{i}.dpt", wn, raw, "#89b4fa")
                for i in range(3)
            ]
            for entry in entries:
                window.spectrum_list.add_entry(entry, select=False, emit_signal=False)
                window._spectrum_states[entry.filepath] = {
                    "wn_crop": wn.copy(),
                    "ab_crop": raw.copy(),
                    "baseline": np.zeros_like(raw),
                    "ab_corrected": raw.copy(),
                    "fit_result": None,
                    "guesses": [PeakGuess(center=3500.0, amplitude=1.0, sigma=70.0, index=0)],
                    "locks": [],
                    "baseline_points": [],
                    "snapshots": [],
                }

            def fake_warning(parent, title, text, *args, **kwargs):
                warnings.append((title, text))
                return QMessageBox.Ok

            QMessageBox.warning = fake_warning

            window._auto_fit()

            self.assertTrue(warnings)
            self.assertEqual(warnings[-1][0], "Auto Fit")
            self.assertIn("첫 번째", warnings[-1][1])
        finally:
            QMessageBox.warning = original_warning
            window.close()

    def test_total_baseline_view_is_rebuilt_after_switching_spectra(self):
        window = MainWindow()
        try:
            wn = np.linspace(1500.0, 4000.0, 251)
            raw_one = 0.01 + 0.02 * np.exp(-((wn - 3350.0) / 200.0) ** 2)
            raw_two = 0.02 + 0.01 * np.exp(-((wn - 2500.0) / 150.0) ** 2)
            entry_one = SpectrumEntry(
                "/tmp/switch-one.dpt", "switch-one.dpt", wn, raw_one, "#89b4fa")
            entry_two = SpectrumEntry(
                "/tmp/switch-two.dpt", "switch-two.dpt", wn, raw_two, "#fab387")
            window._current_entry = entry_one
            window.right_panel.set_wavenumber_range(1500.0, 4000.0)
            window.right_panel.set_mode("Total")
            window.right_panel.btn_edit_bl.setChecked(True)
            window._on_baseline_point_added(1600.0, 999.0)
            window._on_baseline_point_added(3900.0, -999.0)
            corrected_one = window._total_baseline_states[entry_one.filepath]["corrected"].copy()

            window._on_spectrum_selected(entry_two)
            self.assertEqual(window.plot_widget._total_specs[0]["filepath"], entry_two.filepath)

            window._on_spectrum_selected(entry_one)
            self.assertEqual(window.plot_widget._total_specs[0]["filepath"], entry_one.filepath)
            np.testing.assert_allclose(
                window.plot_widget._total_specs[0]["ab"], corrected_one)
            self.assertEqual(len(window.plot_widget._bl_scatter_items), 2)
            self.assertIn("raw", window.plot_widget._items)
            self.assertIn("baseline", window.plot_widget._items)
        finally:
            window.close()

    def test_legacy_oh_corrected_state_is_restored_without_baseline_editor(self):
        window = MainWindow()
        try:
            wn = np.linspace(3000.0, 3990.0, 100)
            raw = 0.01 + wn * 1e-6
            baseline = np.linspace(0.001, 0.002, len(wn))
            corrected = raw - baseline
            entry = SpectrumEntry(
                "/tmp/legacy-oh.dpt", "legacy-oh.dpt", wn, raw, "#89b4fa")
            window._spectrum_states[entry.filepath] = {
                "wn_crop": wn.copy(),
                "ab_crop": raw.copy(),
                "baseline": baseline.copy(),
                "ab_corrected": corrected.copy(),
                "fit_result": None,
                "guesses": [],
                "locks": [],
                "baseline_points": [(3000.0, raw[0]), (3990.0, raw[-1])],
            }

            window._on_spectrum_selected(entry)

            np.testing.assert_allclose(window._ab_corrected, corrected)
            self.assertTrue(window.right_panel._baseline_group.isHidden())
            self.assertEqual(window.plot_widget._bl_scatter_items, [])
        finally:
            window.close()

    def test_spectrum_export_uses_total_corrected_or_raw_for_co_input(self):
        wn = np.linspace(1600.0, 2120.0, 53)
        raw = np.ones_like(wn)
        total_entry = SpectrumEntry(
            "/tmp/export-total.dpt", "export-total.dpt", wn, raw, "#89b4fa")
        raw_entry = SpectrumEntry(
            "/tmp/export-raw.dpt", "export-raw.dpt", wn, raw * 2, "#fab387")
        states = {
            total_entry.filepath: {
                "wn_crop": wn.copy(),
                "ab_crop": raw.copy(),
                "baseline": np.full_like(raw, 0.3),
                "ab_corrected": np.full_like(raw, 0.7),
            }
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            path = str(Path(tmpdir) / "spectra.xlsx")
            export_spectra_excel(
                [total_entry, raw_entry], {}, path, spectrum_states=states)
            workbook = load_workbook(path, data_only=True)
            sheet = workbook["CO Processed"]
            headers = [cell.value for cell in sheet[1]]
            rows = list(sheet.iter_rows(min_row=2, values_only=True))

        self.assertNotIn("Endpoint Source", headers)
        self.assertIn("Input Source", headers)
        source_col = headers.index("Input Source")
        baseline_col = headers.index("Baseline")
        input_col = headers.index("Analysis Input")
        total_row = next(row for row in rows if row[0] == total_entry.name)
        raw_row = next(row for row in rows if row[0] == raw_entry.name)
        self.assertEqual(total_row[source_col], "Total corrected")
        self.assertAlmostEqual(total_row[baseline_col], 0.3)
        self.assertAlmostEqual(total_row[input_col], 0.7)
        self.assertEqual(raw_row[source_col], "Raw")
        self.assertAlmostEqual(raw_row[baseline_col], 0.0)
        self.assertAlmostEqual(raw_row[input_col], 2.0)

    def test_spectrum_export_includes_integrated_areas_sheet(self):
        wn = np.array([1300.0, 1400.0, 1450.0])
        raw = np.ones_like(wn)
        entry = SpectrumEntry(
            "/tmp/export-integral.dpt", "export-integral.dpt", wn, raw, "#89b4fa")
        integrated_records = [
            {
                "session_label": "Sample A",
                "region_name": "NO3",
                "filename": entry.name,
                "potential": -0.1,
                "wn_min": 1300.0,
                "wn_max": 1450.0,
                "area": 0.123456789,
                "source": "corrected",
            }
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            path = str(Path(tmpdir) / "spectra.xlsx")
            info = export_spectra_excel(
                [entry],
                {entry.name: -0.1},
                path,
                integrated_records=integrated_records,
            )
            workbook = load_workbook(path, data_only=True)
            sheet = workbook["Integrated Areas"]
            headers = [cell.value for cell in sheet[1]]
            row = [cell.value for cell in sheet[2]]

        self.assertEqual(info["integral_rows"], 1)
        self.assertEqual(headers, [
            "Session",
            "Region",
            "Spectrum",
            "Potential (V)",
            "Region Min (cm⁻¹)",
            "Region Max (cm⁻¹)",
            "Area",
            "Input Source",
        ])
        self.assertEqual(row[0], "Sample A")
        self.assertEqual(row[1], "NO3")
        self.assertEqual(row[2], entry.name)
        self.assertAlmostEqual(row[3], -0.1)
        self.assertAlmostEqual(row[6], 0.12345679)
        self.assertEqual(row[7], "corrected")

    def test_sio_area_uses_total_corrected_with_user_region(self):
        window = MainWindow()
        try:
            wn = np.linspace(1000.0, 1400.0, 401)
            raw = np.full_like(wn, 10.0)
            corrected = -0.02 * np.exp(-((wn - 1230.0) / 35.0) ** 2)
            entry = SpectrumEntry(
                "/tmp/sio-total-corrected.dpt",
                "sio-total-corrected.dpt",
                wn,
                raw,
                "#89b4fa",
            )
            window.spectrum_list.add_entry(entry)
            window._current_entry = entry
            window.right_panel.set_mode("SiO")
            window.right_panel.spin_wn_min.setValue(1120.0)
            window.right_panel.spin_wn_max.setValue(1280.0)
            window._total_baseline_states[entry.filepath] = {
                "wn": wn.copy(),
                "raw": raw.copy(),
                "baseline": raw - corrected,
                "corrected": corrected.copy(),
            }
            window._on_region_changed(1120.0, 1280.0)

            window._calc_sio_area()

            mask = (wn >= 1120.0) & (wn <= 1280.0)
            trapz = np.trapezoid if hasattr(np, "trapezoid") else np.trapz
            expected = abs(float(trapz(corrected[mask], wn[mask])))
            state = window._sio_states[entry.filepath]
            self.assertAlmostEqual(state["area"], expected)
            self.assertEqual(state["region"], (1120.0, 1280.0))
            self.assertEqual(state["source"], "total")
        finally:
            window.close()

    def test_sio_view_shows_region_handles_without_local_baseline(self):
        window = MainWindow()
        try:
            wn = np.linspace(1000.0, 1400.0, 401)
            raw = np.ones_like(wn)
            corrected = -0.02 * np.exp(-((wn - 1230.0) / 35.0) ** 2)
            entry = SpectrumEntry(
                "/tmp/sio-region-handles.dpt",
                "sio-region-handles.dpt",
                wn,
                raw,
                "#89b4fa",
            )
            window._total_baseline_states[entry.filepath] = {
                "wn": wn.copy(),
                "raw": raw.copy(),
                "baseline": raw - corrected,
                "corrected": corrected.copy(),
            }
            window.plot_widget.show_sio_baseline(wn, raw)
            self.assertTrue(window.plot_widget._ep_lines)
            self.assertIn("bl_SiO", window.plot_widget._items)

            window._apply_sio_view(entry)

            self.assertNotIn("raw", window.plot_widget._items)
            self.assertIn("corrected", window.plot_widget._items)
            self.assertIn("SiO_0", window.plot_widget._ep_lines)
            self.assertIn("SiO_1", window.plot_widget._ep_lines)
            self.assertNotIn("bl_SiO", window.plot_widget._items)
        finally:
            window.close()

    def test_sio_view_keeps_full_corrected_spectrum_outside_handles(self):
        window = MainWindow()
        try:
            wn = np.linspace(1000.0, 1400.0, 401)
            raw = np.ones_like(wn)
            corrected = -0.02 * np.exp(-((wn - 1230.0) / 35.0) ** 2)
            entry = SpectrumEntry(
                "/tmp/sio-full-corrected-display.dpt",
                "sio-full-corrected-display.dpt",
                wn,
                raw,
                "#89b4fa",
            )
            window.spectrum_list.add_entry(entry)
            window._current_entry = entry
            window._total_baseline_states[entry.filepath] = {
                "wn": wn.copy(),
                "raw": raw.copy(),
                "baseline": raw - corrected,
                "corrected": corrected.copy(),
            }
            window.right_panel.set_mode("SiO")
            window.right_panel.set_region_values(1180.0, 1270.0)

            calls = []
            original_show = window.plot_widget.show_highlighted_region

            def capture_highlighted_region(x, y):
                calls.append((np.asarray(x, dtype=float), np.asarray(y, dtype=float)))
                original_show(x, y)

            window.plot_widget.show_highlighted_region = capture_highlighted_region

            window._apply_sio_view(entry)

            self.assertTrue(calls)
            displayed_wn, displayed_corrected = calls[-1]
            np.testing.assert_allclose(displayed_wn, wn)
            np.testing.assert_allclose(displayed_corrected, corrected)

            calls.clear()
            window.plot_widget.update_sio_endpoints(1180.0, 1270.0)
            window._on_sio_endpoint_moved(0, 1180.0)
            window._calc_sio_area()

            self.assertEqual(calls, [])
        finally:
            window.close()

    def test_sio_area_uses_region_handles_over_spinbox_region(self):
        window = MainWindow()
        try:
            wn = np.linspace(1000.0, 1400.0, 401)
            raw = np.zeros_like(wn)
            corrected = -0.02 * np.exp(-((wn - 1230.0) / 35.0) ** 2)
            entry = SpectrumEntry(
                "/tmp/sio-handle-region.dpt",
                "sio-handle-region.dpt",
                wn,
                raw,
                "#89b4fa",
            )
            window.spectrum_list.add_entry(entry)
            window._current_entry = entry
            window.right_panel.set_mode("SiO")
            window.right_panel.spin_wn_min.setValue(1000.0)
            window.right_panel.spin_wn_max.setValue(1400.0)
            window._total_baseline_states[entry.filepath] = {
                "wn": wn.copy(),
                "raw": raw.copy(),
                "baseline": raw - corrected,
                "corrected": corrected.copy(),
            }
            window._apply_sio_view(entry)
            window.plot_widget.update_sio_endpoints(1180.0, 1270.0)

            window._calc_sio_area()

            mask = (wn >= 1180.0) & (wn <= 1270.0)
            trapz = np.trapezoid if hasattr(np, "trapezoid") else np.trapz
            expected = abs(float(trapz(corrected[mask], wn[mask])))
            state = window._sio_states[entry.filepath]
            self.assertAlmostEqual(state["area"], expected)
            self.assertEqual(state["region"], (1180.0, 1270.0))
        finally:
            window.close()


if __name__ == "__main__":
    unittest.main()
